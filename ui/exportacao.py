"""
ui/exportacao.py
================
Funções de exportação de Treeview para CSV e XLSX.
Recebem os dados já extraídos (colunas, linhas) para
não depender diretamente do widget Treeview.
"""

import os
import csv
from tkinter import filedialog, messagebox

from config import OPENPYXL_DISPONIVEL

if OPENPYXL_DISPONIVEL:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment


# ── Extração de dados do Treeview ────────────────────────

def extrair_dados_treeview(tree) -> tuple[list[str], list[list]]:
    """Retorna (colunas, linhas) a partir de um widget Treeview."""
    colunas = list(tree["columns"])
    linhas = [
        [tree.set(k, col) for col in colunas]
        for k in tree.get_children("")
    ]
    return colunas, linhas


# ── Exportação CSV ────────────────────────────────────────

def exportar_csv(
    colunas: list[str],
    linhas: list[list],
    nome_sugerido: str = "exportacao.csv",
    atualizar_rodape=None,
) -> None:
    """Abre diálogo salvar e escreve arquivo CSV (sep=;, UTF-8 BOM)."""
    if not linhas:
        messagebox.showwarning("Exportar CSV", "Nenhum dado para exportar!")
        return

    caminho = filedialog.asksaveasfilename(
        title="Salvar CSV",
        defaultextension=".csv",
        initialfile=nome_sugerido,
        filetypes=[("CSV", "*.csv"), ("Todos", "*.*")],
    )
    if not caminho:
        return

    try:
        with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(colunas)
            writer.writerows(linhas)

        msg = f"✔  Exportado: {os.path.basename(caminho)}  ({len(linhas)} linhas)"
        if atualizar_rodape:
            atualizar_rodape(msg)
    except Exception as e:
        messagebox.showerror("Erro ao exportar CSV", str(e))


# ── Exportação XLSX ───────────────────────────────────────

def exportar_xlsx(
    colunas: list[str],
    linhas: list[list],
    nome_sugerido: str = "exportacao.xlsx",
    titulo_aba: str = "Dados",
    atualizar_rodape=None,
) -> None:
    """Abre diálogo salvar e escreve arquivo XLSX formatado."""
    if not OPENPYXL_DISPONIVEL:
        messagebox.showerror(
            "openpyxl não instalado",
            "Instale com:\n\n   pip install openpyxl\n\ne tente novamente.",
        )
        return

    if not linhas:
        messagebox.showwarning("Exportar XLSX", "Nenhum dado para exportar!")
        return

    caminho = filedialog.asksaveasfilename(
        title="Salvar XLSX",
        defaultextension=".xlsx",
        initialfile=nome_sugerido,
        filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
    )
    if not caminho:
        return

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = titulo_aba[:31]

        fill_h = PatternFill("solid", fgColor="1E3A5F")
        font_h = Font(bold=True, color="FFFFFF", name="Consolas", size=9)

        for ci, col in enumerate(colunas, 1):
            cel = ws.cell(row=1, column=ci, value=col)
            cel.font = font_h
            cel.fill = fill_h
            cel.alignment = Alignment(horizontal="center")

        for ri, linha in enumerate(linhas, 2):
            for ci, val in enumerate(linha, 1):
                ws.cell(row=ri, column=ci, value=val)

        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value or "")) for c in col_cells), default=10
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 2, 50
            )

        wb.save(caminho)
        msg = f"✔  Exportado: {os.path.basename(caminho)}  ({len(linhas)} linhas)"
        if atualizar_rodape:
            atualizar_rodape(msg)
    except Exception as e:
        messagebox.showerror("Erro ao exportar XLSX", str(e))
